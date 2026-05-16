"""
全自动表格复制工具
功能：自动从网页表格复制数据到Excel，支持滚动加载
"""
import time
import pyperclip
import openpyxl
from datetime import datetime
import os
import re

try:
    import pyautogui
    pyautogui.FAILSAFE = False
    pyautogui.PAUSE = 0.01
    HAS_PYAUTOGUI = True
except ImportError:
    HAS_PYAUTOGUI = False
    print("错误：需要安装pyautogui")
    print("pip install pyautogui")


class AutoCopyTool:
    def __init__(self):
        self.copied_count = 0
        self.current_row = 1
        self.output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'output')
        os.makedirs(self.output_dir, exist_ok=True)
        self.init_excel()

        # 配置参数（可根据实际情况调整）
        self.row_height = 25  # 每行高度（像素）
        self.batch_size = 100  # 每批复制数量
        self.scroll_pause = 1.0  # 滚动后等待时间（秒）

    def init_excel(self):
        """初始化Excel文件"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'DataWorks数据_{timestamp}.xlsx'
        self.excel_file = os.path.join(self.output_dir, filename)
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = '数据'

        # 添加表头
        headers = ['序号', '车牌号', '车辆ID', '车架号', '其他数据']
        for col, header in enumerate(headers, 1):
            self.ws.cell(row=1, column=col, value=header)
        self.current_row = 2

        self.save_excel()
        print(f"Excel文件已创建：{self.excel_file}")

    def save_excel(self):
        """保存Excel"""
        self.wb.save(self.excel_file)

    def paste_to_excel(self, data):
        """将复制的数据粘贴到Excel"""
        lines = data.strip().split('\n')
        count = 0

        for line in lines:
            if not line.strip():
                continue

            # 尝试按Tab分割
            if '\t' in line:
                cells = line.split('\t')
            else:
                # 按空格分割，但保留车牌号等含空格的数据
                cells = re.split(r'\s{2,}', line)
                if len(cells) == 1:
                    cells = line.split()

            if cells:
                # 清理每个单元格
                cells = [cell.strip() for cell in cells if cell.strip()]

                # 写入Excel
                for col_idx, cell in enumerate(cells, 1):
                    self.ws.cell(row=self.current_row, column=col_idx, value=cell)

                self.current_row += 1
                count += 1

        self.copied_count += count
        self.save_excel()
        return count

    def get_clipboard_data(self):
        """获取剪贴板数据"""
        try:
            data = pyperclip.paste()
            return data
        except:
            return ""

    def calibrate_position(self):
        """校准位置 - 记录第1行的屏幕坐标"""
        print("\n" + "="*50)
        print("📍 校准位置")
        print("="*50)
        print("请按照以下步骤操作：")
        print("1. 在DataWorks中找到第1行数据")
        print("2. 将鼠标移到第1行的任意位置")
        print("3. 按 Enter 记录位置")
        print("="*50)

        input("\n准备好后按 Enter...")

        # 倒计时
        for i in range(3, 0, -1):
            print(f"{i}秒后记录位置...")
            time.sleep(1)

        # 记录当前位置
        x, y = pyautogui.position()
        self.start_x = x
        self.start_y = y

        print(f"\n✅ 已记录位置：({x}, {y})")

        # 询问行高
        try:
            height = input("\n请输入每行高度（像素，默认25）：").strip()
            if height:
                self.row_height = int(height)
        except:
            pass

        print(f"✅ 行高：{self.row_height} 像素")

    def select_rows_on_screen(self, start_y, end_y):
        """在屏幕上选择指定Y范围的行"""
        x = self.start_x

        try:
            # 点击起始位置
            pyautogui.click(x, start_y)
            time.sleep(0.1)

            # 按住Shift点击结束位置
            pyautogui.keyDown('shift')
            pyautogui.click(x, end_y)
            pyautogui.keyUp('shift')
            time.sleep(0.2)

            # 复制
            pyautogui.hotkey('ctrl', 'c')
            time.sleep(0.3)

            return True
        except Exception as e:
            print(f"❌ 选择失败：{e}")
            return False

    def scroll_down(self, clicks=5):
        """向下滚动表格"""
        print(f"⬇️ 滚动表格...")
        pyautogui.scroll(-clicks, self.start_x, 400)  # 向下滚动
        time.sleep(self.scroll_pause)

    def scroll_up_to_top(self):
        """滚动到顶部"""
        print("⬆️ 滚动到顶部...")
        # 按Home键回到顶部
        pyautogui.press('home')
        time.sleep(1)

    def copy_batch_from_screen(self, visible_rows):
        """从屏幕上复制可见的行"""
        # 计算屏幕上可见行的Y范围
        start_y = self.start_y
        end_y = self.start_y + (visible_rows - 1) * self.row_height

        print(f"📍 选择屏幕上的行：Y={start_y} 到 Y={end_y}")

        if not self.select_rows_on_screen(start_y, end_y):
            return 0

        # 获取剪贴板数据
        data = self.get_clipboard_data()
        if not data:
            print("❌ 复制失败，剪贴板为空")
            return 0

        # 粘贴到Excel
        count = self.paste_to_excel(data)
        print(f"✅ 成功复制 {count} 条数据")

        return count

    def auto_copy_by_total(self, total_count):
        """根据总数自动复制（支持滚动）"""
        print(f"\n🚀 开始自动复制，共 {total_count} 条")
        print(f"   每批 {self.batch_size} 条")

        # 先滚动到顶部
        self.scroll_up_to_top()
        time.sleep(1)

        copied = 0
        batch_num = 0

        while copied < total_count:
            batch_num += 1
            remaining = total_count - copied
            current_batch = min(self.batch_size, remaining)

            print(f"\n--- 第 {batch_num} 批 ---")
            print(f"   需要复制：{current_batch} 条")

            # 计算当前屏幕可见的行数
            visible_rows = min(current_batch, self.batch_size)

            # 从屏幕复制
            count = self.copy_batch_from_screen(visible_rows)

            if count == 0:
                print("❌ 复制失败，停止")
                break

            copied += count
            print(f"📊 已复制总数：{copied}/{total_count}")

            # 如果还有更多数据，滚动
            if copied < total_count:
                # 向下滚动，让下一批数据可见
                self.scroll_down(self.batch_size)
                print(f"⏸️ 等待 {self.scroll_pause} 秒加载...")
                time.sleep(self.scroll_pause)

        print(f"\n🎉 完成！共复制 {copied} 条数据")
        return copied

    def auto_copy_range(self, start_row, end_row):
        """自动复制指定行号范围（需要滚动）"""
        total = end_row - start_row + 1
        print(f"\n🚀 自动复制：第 {start_row} 行 到 第 {end_row} 行，共 {total} 条")

        # 滚动到顶部
        self.scroll_up_to_top()
        time.sleep(1)

        copied = 0
        current_row = start_row

        while current_row <= end_row:
            # 计算当前批次
            remaining = end_row - current_row + 1
            current_batch = min(self.batch_size, remaining)

            print(f"\n--- 当前行：{current_row} ---")

            # 从屏幕复制
            count = self.copy_batch_from_screen(current_batch)

            if count == 0:
                print("❌ 复制失败，停止")
                break

            copied += count
            current_row += count
            print(f"📊 已复制：{copied} 条，下一行：{current_row}")

            # 如果还有更多数据，滚动
            if current_row <= end_row:
                self.scroll_down(self.batch_size)
                print(f"⏸️ 等待加载...")
                time.sleep(self.scroll_pause)

        print(f"\n🎉 完成！共复制 {copied} 条数据")
        return copied

    def manual_copy_mode(self):
        """手动复制模式"""
        print("\n" + "="*50)
        print("📋 手动复制模式")
        print("="*50)
        print("操作步骤：")
        print("1. 在DataWorks中选择要复制的行（最多100行）")
        print("2. Ctrl+C 复制")
        print("3. 在此窗口按 Enter 粘贴到Excel")
        print("4. 输入 'quit' 退出程序")
        print("="*50)

        while True:
            try:
                cmd = input("\n按 Enter 粘贴，输入 quit 退出：").strip()
                if cmd.lower() == 'quit':
                    break

                data = self.get_clipboard_data()
                if data:
                    count = self.paste_to_excel(data)
                    print(f"✅ 成功粘贴 {count} 条数据")
                    print(f"📊 已复制总数：{self.copied_count} 条")
                else:
                    print("⚠️ 剪贴板为空，请先复制数据")

            except KeyboardInterrupt:
                break
            except Exception as e:
                print(f"❌ 错误：{e}")

    def auto_copy_mode(self):
        """全自动复制模式"""
        print("\n" + "="*50)
        print("🤖 全自动复制模式")
        print("="*50)

        # 校准位置
        self.calibrate_position()

        while True:
            print("\n选择操作：")
            print("1. 复制指定行号范围")
            print("2. 根据总数复制")
            print("3. 返回主菜单")

            choice = input("\n请选择：").strip()

            if choice == '1':
                try:
                    start = int(input("起始行号："))
                    end = int(input("结束行号："))
                    self.auto_copy_range(start, end)
                except ValueError:
                    print("❌ 请输入有效的数字")

            elif choice == '2':
                try:
                    total = int(input("总行数："))
                    self.auto_copy_by_total(total)
                except ValueError:
                    print("❌ 请输入有效的数字")

            elif choice == '3':
                break

    def show_summary(self):
        """显示汇总信息"""
        print("\n" + "="*50)
        print("📊 汇总信息")
        print("="*50)
        print(f"已复制总数：{self.copied_count} 条")
        print(f"Excel文件：{self.excel_file}")
        print("="*50)

    def run(self):
        """主程序"""
        print("="*60)
        print("🔧 全自动表格复制工具")
        print("="*60)
        print("功能：自动从网页表格复制数据到Excel")
        print("     支持滚动加载，批量复制")
        print("="*60)

        if not HAS_PYAUTOGUI:
            print("\n❌ 未检测到pyautogui，仅支持手动模式")
            print("安装方法：pip install pyautogui")

        while True:
            print("\n选择模式：")
            print("1. 全自动复制（推荐）")
            print("2. 手动复制")
            print("3. 查看统计")
            print("4. 退出")

            choice = input("\n请输入选项：").strip()

            if choice == '1' and HAS_PYAUTOGUI:
                self.auto_copy_mode()
            elif choice == '2' or (choice == '1' and not HAS_PYAUTOGUI):
                self.manual_copy_mode()
            elif choice == '3':
                self.show_summary()
            elif choice == '4':
                break
            else:
                print("❌ 无效选项")

        print(f"\n👋 程序退出，Excel文件已保存在：{self.excel_file}")


if __name__ == "__main__":
    tool = AutoCopyTool()
    tool.run()
