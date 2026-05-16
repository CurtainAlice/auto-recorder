"""
智能表格复制工具 - OCR版
功能：OCR识别网页表格序号，自动复制指定范围的数据到Excel
"""
import time
import pyperclip
import openpyxl
from datetime import datetime
import os
import sys
import re
import json

# 尝试导入OCR库
try:
    import pytesseract
    from PIL import Image, ImageGrab
    HAS_OCR = True
except ImportError:
    HAS_OCR = False

# 尝试导入GUI库
try:
    import pyautogui
    HAS_PYAUTOGUI = True
except ImportError:
    HAS_PYAUTOGUI = False


class SmartCopyTool:
    def __init__(self):
        self.copied_count = 0
        self.current_row = 1
        self.output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'output')
        os.makedirs(self.output_dir, exist_ok=True)

        # 配置Tesseract路径（Windows）
        if HAS_OCR and sys.platform == 'win32':
            possible_paths = [
                r"C:\Program Files\Tesseract-OCR\tesseract.exe",
                r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
            ]
            for path in possible_paths:
                if os.path.exists(path):
                    pytesseract.pytesseract.tesseract_cmd = path
                    break

        self.init_excel()

    def init_excel(self):
        """初始化Excel文件"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'DataWorks数据_{timestamp}.xlsx'
        self.excel_file = os.path.join(self.output_dir, filename)
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = '数据'

        # 添加表头
        headers = ['序号', '车牌号', '车辆ID', '车架号', '其他数据']
        for col, header in enumerate(headers, 1):
            self.ws.cell(row=1, column=col, value=header)
        self.current_row = 2

        self.save_excel()
        print(f"Excel文件已创建：{self.excel_file}")

    def save_excel(self):
        """保存Excel"""
        self.wb.save(self.excel_file)

    def paste_to_excel(self, data):
        """将复制的数据粘贴到Excel"""
        lines = data.strip().split('\n')
        count = 0

        for line in lines:
            if not line.strip():
                continue

            # 尝试按Tab分割
            if '\t' in line:
                cells = line.split('\t')
            else:
                # 按空格分割，但保留车牌号等含空格的数据
                cells = re.split(r'\s{2,}', line)
                if len(cells) == 1:
                    cells = line.split()

            if cells:
                # 清理每个单元格
                cells = [cell.strip() for cell in cells if cell.strip()]

                # 写入Excel
                for col_idx, cell in enumerate(cells, 1):
                    self.ws.cell(row=self.current_row, column=col_idx, value=cell)

                self.current_row += 1
                count += 1

        self.copied_count += count
        self.save_excel()
        return count

    def get_clipboard_data(self):
        """获取剪贴板数据"""
        try:
            data = pyperclip.paste()
            return data
        except:
            return ""

    def ocr_screen_region(self, x1, y1, x2, y2):
        """OCR识别屏幕区域"""
        if not HAS_OCR:
            return ""

        try:
            screenshot = ImageGrab.grab(bbox=(x1, y1, x2, y2))
            text = pytesseract.image_to_string(screenshot, lang='chi_sim+eng')
            return text
        except Exception as e:
            print(f"OCR识别失败：{e}")
            return ""

    def find_row_numbers_on_screen(self):
        """识别屏幕上的所有行号"""
        if not HAS_OCR:
            print("❌ OCR功能未安装")
            return []

        print("\n🔍 正在识别屏幕上的行号...")
        screen_width, screen_height = pyautogui.size()

        # 截取屏幕左侧区域（序号列通常在左侧）
        region_width = min(200, screen_width // 4)
        text = self.ocr_screen_region(0, 0, region_width, screen_height)

        # 提取数字（行号）
        numbers = re.findall(r'\b\d+\b', text)
        row_numbers = []
        for num in numbers:
            try:
                row_numbers.append(int(num))
            except:
                pass

        return sorted(set(row_numbers))

    def auto_copy_by_row_range(self, start_row, end_row):
        """根据行号范围自动复制"""
        if not HAS_OCR or not HAS_PYAUTOGUI:
            print("❌ 自动复制需要OCR和GUI库")
            print("请安装：pip install pytesseract Pillow pyautogui")
            return

        print(f"\n🚀 自动复制：第 {start_row} 行 到第 {end_row} 行")
        print("="*50)

        # 等待用户准备好
        input("\n请确保DataWorks窗口在前台，然后按 Enter 继续...")

        # 识别行号位置
        screen_width, screen_height = pyautogui.size()
        region_width = min(200, screen_width // 4)

        print("\n🔍 正在识别行号位置...")

        # 截取屏幕进行OCR
        screenshot = ImageGrab.grab(bbox=(0, 0, region_width, screen_height))
        text = pytesseract.image_to_string(screenshot, lang='chi_sim+eng')

        # 解析行号和位置
        lines = text.split('\n')
        row_positions = {}

        for i, line in enumerate(lines):
            numbers = re.findall(r'\b\d+\b', line)
            for num in numbers:
                try:
                    row_num = int(num)
                    if row_num == start_row:
                        y_pos = int((i / len(lines)) * screen_height)
                        row_positions[row_num] = y_pos
                except:
                    pass

        if start_row not in row_positions:
            print(f"⚠️ 未找到行号 {start_row}，请手动定位")
            return

        start_y = row_positions[start_row]
        x_pos = region_width // 2  # 序号列中间位置

        print(f"\n📍 找到起始行 {start_row}，位置：({x_pos}, {start_y})")

        # 估算结束行位置
        row_height = 25  # 假设每行25像素
        if end_row in row_positions:
            end_y = row_positions[end_row]
        else:
            end_y = start_y + (end_row - start_row) * row_height
            print(f"⚠️ 使用估算位置：{end_y}")

        print("3秒后开始操作...")
        time.sleep(3)

        try:
            # 点击起始行
            pyautogui.click(x_pos, start_y)
            time.sleep(0.2)

            # 按住Shift点击结束行
            pyautogui.keyDown('shift')
            pyautogui.click(x_pos, end_y)
            pyautogui.keyUp('shift')
            time.sleep(0.3)

            # 复制
            pyautogui.hotkey('ctrl', 'c')
            time.sleep(0.5)

            # 获取剪贴板内容
            data = self.get_clipboard_data()
            if data:
                count = self.paste_to_excel(data)
                print(f"\n✅ 成功复制 {count} 条数据")
                print(f"📊 已复制总数：{self.copied_count} 条")
            else:
                print("❌ 复制失败，剪贴板为空")

        except Exception as e:
            print(f"❌ 操作失败：{e}")

    def manual_copy_mode(self):
        """手动复制模式"""
        print("\n" + "="*50)
        print("📋 手动复制模式")
        print("="*50)
        print("操作步骤：")
        print("1. 在DataWorks中选择要复制的行（最多100行）")
        print("2. Ctrl+C 复制")
        print("3. 在此窗口按 Enter 粘贴到Excel")
        print("4. 输入 'quit' 退出程序")
        print("="*50)

        while True:
            try:
                cmd = input("\n按 Enter 粘贴，输入 quit 退出：").strip()
                if cmd.lower() == 'quit':
                    break

                data = self.get_clipboard_data()
                if data:
                    count = self.paste_to_excel(data)
                    print(f"✅ 成功粘贴 {count} 条数据")
                    print(f"📊 已复制总数：{self.copied_count} 条")
                else:
                    print("⚠️ 剪贴板为空，请先复制数据")

            except KeyboardInterrupt:
                break
            except Exception as e:
                print(f"❌ 错误：{e}")

        self.show_summary()

    def auto_copy_mode(self):
        """自动复制模式"""
        if not HAS_OCR:
            print("❌ OCR功能未安装")
            print("请安装：pip install pytesseract Pillow")
            print("还需要安装Tesseract OCR引擎")
            return

        print("\n" + "="*50)
        print("🤖 自动复制模式")
        print("="*50)

        # 选择操作方式
        print("\n选择操作方式：")
        print("1. 识别屏幕上所有行号")
        print("2. 手动输入行号范围")

        choice = input("\n请选择 (1-2)：").strip()

        if choice == '1':
            row_numbers = self.find_row_numbers_on_screen()
            if row_numbers:
                print(f"\n识别到的行号：{row_numbers}")
                start = int(input("起始行号："))
                end = int(input("结束行号："))
                self.auto_copy_by_row_range(start, end)
            else:
                print("❌ 未识别到行号")
        elif choice == '2':
            try:
                start = int(input("起始行号："))
                end = int(input("结束行号："))
                if end - start + 1 > 100:
                    print("⚠️ 超过100条限制，自动截断到100条")
                    end = start + 99
                self.auto_copy_by_row_range(start, end)
            except ValueError:
                print("❌ 请输入有效的数字")

    def show_summary(self):
        """显示汇总信息"""
        print("\n" + "="*50)
        print("📊 汇总信息")
        print("="*50)
        print(f"已复制总数：{self.copied_count} 条")
        print(f"Excel文件：{self.excel_file}")
        print("="*50)

    def run(self):
        """主程序"""
        print("="*60)
        print("🔧 智能表格复制工具 - OCR版")
        print("="*60)
        print("功能：OCR识别网页表格序号，自动复制数据到Excel")
        print("="*60)

        if HAS_OCR:
            print("✅ OCR功能已启用")
        else:
            print("⚠️ OCR功能未安装，仅支持手动模式")

        while True:
            print("\n选择模式：")
            print("1. 手动复制模式")
            if HAS_OCR:
                print("2. 自动复制模式（OCR识别）")
            print("3. 查看统计")
            print("4. 退出")

            choice = input("\n请输入选项：").strip()

            if choice == '1':
                self.manual_copy_mode()
            elif choice == '2' and HAS_OCR:
                self.auto_copy_mode()
            elif choice == '3':
                self.show_summary()
            elif choice == '4':
                break
            else:
                print("❌ 无效选项")

        print(f"\n👋 程序退出，Excel文件已保存在：{self.excel_file}")


if __name__ == "__main__":
    tool = SmartCopyTool()
    tool.run()
